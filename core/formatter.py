# core/formatter.py
import os
from typing import List, Optional, Dict
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from core.validators import CheckConfig

def write_styled_excel(
    merged_df: pd.DataFrame,
    metadata: Dict,
    output_path: str,
    title_columns: List[Optional[str]],
    check_config: CheckConfig
) -> None:
    # 1. Save raw DataFrame
    merged_df.to_excel(output_path, index=False, header=True)

    # 2. Load workbook & sheet
    wb = load_workbook(output_path, data_only=False, rich_text=True)
    ws = wb.active

    # 3. Reorder columns
    reorder_columns(ws)

    # 4. Apply fills, duplicates, instance/case, hyperlinks
    apply_formatting_and_hyperlinks(ws, metadata, merged_df, check_config)

    # 5. Title-match rich-text coloring ("True"/"False")
    apply_title_match_highlighting(ws, merged_df)

    # 6. Title-to-title rich-text diffing
    apply_title_highlighting(ws, merged_df, title_columns)


    # 7. Save & close
    wb.save(output_path)
    wb.close()


def reorder_columns(ws: Worksheet) -> None:
    original_headers = [cell.value for cell in ws[1]]
    data = list(ws.iter_rows(min_row=2, values_only=True))

    title_cols = [c for c in ["title_excel1", "title_excel2", "title_excel3"] if c in original_headers]
    final_cols = [c for c in ["title_match", "Comments_1", "Instance", "Case","Remerged","Duplicate"] if c in original_headers]

    # add your internal columns to excluded so they get dropped:
    excluded = set(title_cols) | set(final_cols) | {"common_ref", "original_row_index", "original_row_index_3", "original_row_index_2"}

    data_cols = [c for c in original_headers if c not in excluded]
    new_order = (
        data_cols
        + (["common_ref"] if "common_ref" in original_headers else [])
        + title_cols
        + final_cols
    )

    ws.delete_cols(1, ws.max_column)
    for idx, header in enumerate(new_order, start=1):
        ws.cell(row=1, column=idx, value=header)

    header_index = {h: i for i, h in enumerate(original_headers)}
    for row_i, row_vals in enumerate(data, start=2):
        for col_i, header in enumerate(new_order, start=1):
            orig_idx = header_index.get(header)
            ws.cell(row=row_i, column=col_i, value=row_vals[orig_idx] if orig_idx is not None else None)


def apply_formatting_and_hyperlinks(
    ws: Worksheet,
    metadata: Dict,
    merged_df: pd.DataFrame,
    config: CheckConfig
) -> None:
    headers = [cell.value for cell in ws[1]]
    col_idx = {h: i+1 for i, h in enumerate(headers)}

    dup_font = Font(bold=True, color="FF3300")
    has3 = 'number_3' in merged_df.columns

    fills = {
        (True, False, False): PatternFill("solid", fgColor="CC99FF"),
        (False, True, False): PatternFill("solid", fgColor="FFCC66"),
        (False, False, True): PatternFill("solid", fgColor="AACCFF"),
        (True, True, False): PatternFill("solid", fgColor="FFC0CB"),
        (True, False, True): PatternFill("solid", fgColor="90EE90"),
        (False, True, True): PatternFill("solid", fgColor="FFD700"),
    } if has3 else {
        (True, False): PatternFill("solid", fgColor="CC99FF"),
        (False, True): PatternFill("solid", fgColor="FFCC66"),
    }

    dup_cols = ['number_1', 'number_2', 'common_ref'] + (['number_3'] if has3 else [])
    dup_sets = {col: set(merged_df[col][merged_df[col].duplicated()]) for col in dup_cols}

    # Add the Case/Instance column
    inst_col = ws.max_column + 1
    inst_header = "Case" if has3 else "Instance"
    ws.cell(row=1, column=inst_col, value=inst_header)

    # --- Full presence → text mapping ---
    if has3:
        instance_mapping = {
            (True, False, False): "PDF is provided but not listed in LOD",
            (False, True, False): "LOD_2 only",
            (False, False, True): "LOD_3 only",
            (True, True, False): "PDF is provided and number_2",
            (True, False, True): "PDF is provided and number_3",
            (False, True, True): "No PDF but found in LOD_2 and LOD_3",
            (True, True, True): ""   # all three present → blank
        }
    else:
        instance_mapping = {
            (True, False): "PDF is provided but not listed in LOD",
            (False, True): "LOD_2",
            (True, True): ""         # both present → blank
        }
    # -------------------------------------

    # Iterate each row
    for i, _ in merged_df.iterrows():
        r = i + 2

        # Build presence tuple
        if has3:
            presence = (
                bool(ws.cell(r, col_idx['number_1']).value),
                bool(ws.cell(r, col_idx['number_2']).value),
                bool(ws.cell(r, col_idx['number_3']).value),
            )
            number_cols = ['number_1', 'number_2', 'number_3']
        else:
            presence = (
                bool(ws.cell(r, col_idx['number_1']).value),
                bool(ws.cell(r, col_idx['number_2']).value),
            )
            number_cols = ['number_1', 'number_2']

        # Apply fills
        if presence in fills:
            fill = fills[presence]
            for colname, present in zip(number_cols, presence):
                if present:
                    ws.cell(r, col_idx[colname]).fill = fill
            if 'common_ref' in col_idx:
                ws.cell(r, col_idx['common_ref']).fill = fill

        # Apply duplicate-fonts
        for colname, dset in dup_sets.items():
            if colname in col_idx:
                val = ws.cell(r, col_idx[colname]).value
                if val in dset:
                    ws.cell(r, col_idx[colname]).font = dup_font

        # Set Case/Instance text
        ws.cell(r, inst_col).value = instance_mapping.get(presence, "None")

    # Re-insert hyperlinks
    for orig_idx, cols in metadata.get('hyperlinks', {}).items():
        match = merged_df[merged_df.get('original_row_index', -1) == orig_idx]
        if not match.empty:
            r = match.index[0] + 2
            for colname, url in cols.items():
                if colname in col_idx:
                    cell = ws.cell(r, col_idx[colname])
                    cell.hyperlink = url
                    cell.style = "Hyperlink"

    # Highlight rows failing checks
    light_red = PatternFill("solid", fgColor="FFCCCC")
    checks = [
        (config.status_column, config.status_value),
        (config.project_column, config.project_value)
    ] + (config.custom_checks or [])

    for colname, expected in checks:
        if not colname:
            continue
        cidx = col_idx.get(colname)
        if not cidx:
            continue
        for i, row in merged_df.iterrows():
            actual = row.get(colname, "")
            if pd.notna(row.get('number_1')) and str(actual).strip().lower() != str(expected).strip().lower():
                ws.cell(i+2, cidx).fill = light_red

    # 5️⃣ Highlight filename mismatches
    highlight_filename_mismatches(ws, merged_df, col_idx, config)

    # 6️⃣ Highlight any rows that were re‐merged by filename
    if "Remerged" in merged_df.columns and "common_ref" in col_idx:
        cyan = PatternFill("solid", fgColor="00FFFF")
        for i, row in merged_df.iterrows():
            if row.get("Remerged", False):
                r = i + 2
                ws.cell(r, col_idx["common_ref"]).fill = cyan

def get_ws_column_index(ws: Worksheet, header_name: str) -> Optional[int]:
    for cell in ws[1]:
        if isinstance(cell.value, str) and cell.value.strip().lower() == header_name.strip().lower():
            return cell.column
    return None


def apply_title_match_highlighting(ws: Worksheet, merged_df: pd.DataFrame) -> None:
    col_idx = get_ws_column_index(ws, "title_match")
    if not col_idx:
        return

    for i, row in merged_df.iterrows():
        text = str(row.get("title_match", ""))
        rich = CellRichText()
        for j, token in enumerate(text.split(",")):
            tok = token.strip()
            font = InlineFont(rFont="Calibri", sz=11)
            if tok.lower() == "true":
                font.color = "008000"
            elif tok.lower() == "false":
                font.color = "FF0000"
            rich.append(TextBlock(font, tok))
            if j < len(text.split(",")) - 1:
                sep = InlineFont(rFont="Calibri", sz=11, color="000000")
                rich.append(TextBlock(sep, ", "))
        ws.cell(i+2, col_idx).value = rich


def tokenize_with_indices(text: str):
    return [(m.group(), m.start())
            for m in re.finditer(r'[A-Za-z0-9]+|[_–—-]|[^\w\s]', text)]


def dp_align_tokens(tokens1, tokens2):
    from difflib import SequenceMatcher

    s1 = [t[0] for t in tokens1]
    s2 = [t[0] for t in tokens2]
    n, m = len(s1), len(s2)

    # cost matrix & backpointers
    dp = [[0] * (m + 1) for _ in range(n + 1)]
    bt = [[None] * (m + 1) for _ in range(n + 1)]

    # initialize
    for i in range(1, n + 1):
        dp[i][0] = i
        bt[i][0] = "DEL"
    for j in range(1, m + 1):
        dp[0][j] = j
        bt[0][j] = "INS"

    # fill
    for i in range(1, n + 1):
        for j in range(1, m + 1):
            a, b = s1[i - 1], s2[j - 1]
            if a == b:
                cost_sub, flag = dp[i - 1][j - 1], "EXACT"
            elif a.lower() == b.lower():
                cost_sub, flag = dp[i - 1][j - 1] + 0.5, "CASE_ONLY"
            else:
                sim = SequenceMatcher(None, a, b).ratio()
                if sim >= 0.8:
                    cost_sub, flag = dp[i - 1][j - 1] + (1 - sim), "CHAR_LEVEL"
                else:
                    cost_sub, flag = float("inf"), None

            cost_del = dp[i - 1][j] + 1
            cost_ins = dp[i][j - 1] + 1

            best = min(cost_sub, cost_del, cost_ins)
            dp[i][j] = best

            if best == cost_sub:
                bt[i][j] = flag
            elif best == cost_del:
                bt[i][j] = "DEL"
            else:
                bt[i][j] = "INS"

    # backtrack
    aligned1, aligned2, flags = [], [], []
    i, j = n, m
    while i > 0 or j > 0:
        move = bt[i][j]
        if move in ("EXACT", "CASE_ONLY", "CHAR_LEVEL"):
            aligned1.append(tokens1[i - 1])
            aligned2.append(tokens2[j - 1])
            flags.append(move)
            i, j = i - 1, j - 1
        elif move == "DEL":
            aligned1.append(tokens1[i - 1])
            aligned2.append((None, None))
            flags.append("DEL")
            i -= 1
        else:  # INS
            aligned1.append((None, None))
            aligned2.append(tokens2[j - 1])
            flags.append("INS")
            j -= 1

    return aligned1[::-1], aligned2[::-1], flags[::-1]

def create_rich_text(original, aligned1, aligned2, flags):
    """
    Build a CellRichText so that:
      - EXACT → black
      - CASE_ONLY → gray
      - CHAR_LEVEL → per-char: match=black, diff=orange
      - DEL/INS → red
    """
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    rich = CellRichText()
    idx = 0

    for (tok1, pos), (tok2, _), flag in zip(aligned1, aligned2, flags):
        if tok1 is None:
            continue

        # write any intervening literal text
        if pos > idx:
            rich.append(original[idx:pos])

        # choose style by flag
        if flag == "EXACT":
            font = InlineFont(rFont="Calibri", sz=11, color="000000")
            rich.append(TextBlock(font, tok1))

        elif flag == "CASE_ONLY":
            font = InlineFont(rFont="Calibri", sz=11, color="808080")
            rich.append(TextBlock(font, tok1))

        elif flag == "CHAR_LEVEL":
            # compare char by char
            for c_idx, ch in enumerate(tok1):
                other_ch = tok2[c_idx] if tok2 and c_idx < len(tok2) else None
                color = "000000" if ch == other_ch else "FFA500"
                font = InlineFont(rFont="Calibri", sz=11, color=color)
                rich.append(TextBlock(font, ch))

        elif flag == "DEL" or flag == "INS":
            font = InlineFont(rFont="Calibri", sz=11, color="FF0000")
            rich.append(TextBlock(font, tok1))

        else:
            # fallback
            font = InlineFont(rFont="Calibri", sz=11, color="000000")
            rich.append(TextBlock(font, tok1))

        idx = pos + len(tok1)

    # any trailing text
    if idx < len(original):
        rich.append(original[idx:])

    return rich

def highlight_filename_mismatches(
    ws,
    merged_df: pd.DataFrame,
    col_idx: Dict[str,int],
    config: CheckConfig
) -> None:
    """
    After hyperlinks have been applied, for each row where filename doesn't
    start with number_1, split both on '-' and color only the mismatched
    characters in red & bold, leaving matches black. Then apply a yellow fill
    to the filename cell only, and update the number_1 cell with the same
    per-character diff (no fill).
    """
    filename_col = config.filename_column
    if not filename_col or filename_col not in col_idx:
        return

    fidx   = col_idx[filename_col]
    nidx   = col_idx.get("number_1")
    yellow = PatternFill("solid", fgColor="FFFF99")

    for i, row in merged_df.iterrows():
        num = str(row.get("number_1", "")).strip()
        fn  = str(row.get(filename_col, "")).strip()

        # skip if missing either
        if not num or not fn:
            continue

        # strip off any extension (".pdf", ".xlsm", etc.)
        fn_base, ext = os.path.splitext(fn)

        # if filename truly begins with the drawing number (case-insensitive), nothing to highlight
        if fn_base.lower().startswith(num.lower()):
            continue

        # Tokenize on hyphens
        num_tokens  = num.split("-")
        file_tokens = fn_base.split("-")
        compare_len = len(num_tokens)
        base_tokens = file_tokens[:compare_len]
        extra_part  = ""
        if len(file_tokens) > compare_len:
            extra_part = "-" + "-".join(file_tokens[compare_len:])

        # Helper to build fonts
        def font(color, bold=False):
            return InlineFont(rFont="Calibri", sz=11, color=color, b=bold)

        # Build rich text containers
        rich_fn  = CellRichText()
        rich_num = CellRichText()

        # Compare token by token
        for t in range(compare_len):
            ft = base_tokens[t] if t < len(base_tokens) else ""
            nt = num_tokens[t]

            if ft == nt:
                # perfect token match
                rich_fn.append(TextBlock(font("000000"), ft))
                rich_num.append(TextBlock(font("000000"), nt))
            else:
                # char-by-char diff
                common_len = min(len(ft), len(nt))
                for c in range(common_len):
                    fch, nch = ft[c], nt[c]
                    if fch == nch:
                        rich_fn.append(TextBlock(font("000000"), fch))
                        rich_num.append(TextBlock(font("000000"), nch))
                    else:
                        rich_fn.append(TextBlock(font("FF0000", bold=True), fch))
                        rich_num.append(TextBlock(font("FF0000", bold=True), nch))
                # any extra in filename token
                for ch in ft[common_len:]:
                    rich_fn.append(TextBlock(font("FF0000", bold=True), ch))
                # any extra in number_1 token
                for ch in nt[common_len:]:
                    rich_num.append(TextBlock(font("FF0000", bold=True), ch))

            # re-insert hyphen separator if not last
            if t < compare_len - 1:
                sep = TextBlock(font("000000"), "-")
                rich_fn.append(sep)
                rich_num.append(sep)

        # append any “extra” suffix (re-join hyphen + rest)
        if extra_part:
            rich_fn.append(TextBlock(font("000000"), extra_part))

        # re-append extension on filename side
        if ext:
            rich_fn.append(TextBlock(font("000000"), ext))

        # write back into worksheet
        row_idx = i + 2  # account for header
        # Filename cell: rich text + yellow fill
        cell_fn = ws.cell(row=row_idx, column=fidx)
        cell_fn.value = rich_fn
        cell_fn.fill  = yellow

        # number_1 cell: just the rich text (no fill)
        if nidx:
            cell_num = ws.cell(row=row_idx, column=nidx)
            cell_num.value = rich_num

def apply_title_highlighting(
    ws: Worksheet,
    merged_df: pd.DataFrame,
    title_columns: List[Optional[str]]
) -> None:
    from openpyxl.cell.rich_text import CellRichText

    # baseline is first title
    baseline = title_columns[0] if title_columns else None
    if not baseline:
        return

    base_idx = get_ws_column_index(ws, baseline)
    if base_idx is None:
        return

    # for each other title
    for other in title_columns[1:]:
        if not other:
            continue
        other_idx = get_ws_column_index(ws, other)
        if other_idx is None:
            continue

        for i, row in merged_df.iterrows():
            excel_row = i + 2
            t1 = str(row.get(baseline, ""))
            t2 = str(row.get(other, ""))

            # tokenize
            tok1 = tokenize_with_indices(t1)
            tok2 = tokenize_with_indices(t2)

            # align
            aligned1, aligned2, flags = dp_align_tokens(tok1, tok2)

            # build rich text
            rich1 = create_rich_text(t1, aligned1, aligned2, flags)
            rich2 = create_rich_text(t2, aligned2, aligned1, flags)

            # write back
            ws.cell(row=excel_row, column=base_idx).value = rich1
            ws.cell(row=excel_row, column=other_idx).value = rich2