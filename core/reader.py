"""
Reader module: loads Excel files into DataFrames and extracts metadata (hyperlinks, original row indices).
"""
from typing import List, Tuple, Dict
import pandas as pd
from openpyxl import load_workbook


def extract_hyperlinks(path: str) -> Dict[int, Dict[str, str]]:
    """
    Extracts hyperlinks from the given Excel file.

    Returns a dict mapping original row numbers to a dict of {column_header: hyperlink_target}.
    """
    wb = load_workbook(path, data_only=False)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

    hyperlinks: Dict[int, Dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row_idx = row[0].row
        row_links: Dict[str, str] = {}
        for idx, cell in enumerate(row):
            if cell.hyperlink:
                col_name = headers[idx]
                row_links[col_name] = cell.hyperlink.target
        if row_links:
            hyperlinks[row_idx] = row_links
    wb.close()
    return hyperlinks


def extract_original_row_indices(path: str) -> List[int]:
    """
    Returns a list of Excel row numbers for all non-blank data rows (excluding header).
    """
    wb = load_workbook(path, data_only=False)
    ws = wb.active
    original_indices: List[int] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        values = [cell.value for cell in row]
        if any(v is not None and str(v).strip() != "" for v in values):
            original_indices.append(row[0].row)
    wb.close()
    return original_indices



def read_excels(
    paths: List[str],
    ref_columns: List[str]
) -> Tuple[List[pd.DataFrame], Dict]:
    """
    Read each Excel file into a DataFrame, extract original row indices and hyperlinks.

    Args:
        paths: list of Excel file paths.
        ref_columns: list of reference‐column names (only used for key‐existence checks here).

    Returns:
        dfs: List of DataFrames, each with a new 'original_row_index' column.
        metadata: {
            "hyperlinks": {
                original_row_number: { column_header: url, ... },
                ...
            }
        }
    """
    dfs: List[pd.DataFrame] = []
    metadata: Dict = {"hyperlinks": {}}

    for path, ref_col in zip(paths, ref_columns):
        # 1) Open with openpyxl to extract hyperlinks & row indices
        wb = load_workbook(path, data_only=False)
        ws = wb.active

        # Collect original row numbers and hyperlinks per row
        row_indices: List[int] = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            # We consider a row “present” if any cell has a value
            if any(cell.value not in (None, "") for cell in row):
                orig_row = row[0].row
                row_indices.append(orig_row)

                # Gather any hyperlinks in this row
                link_map: Dict[str, str] = {}
                for cell in row:
                    if cell.hyperlink:
                        # header is in row 1 of same column
                        header = ws.cell(row=1, column=cell.column).value
                        link_map[header] = cell.hyperlink.target
                if link_map:
                    # If multiple files share the same row num, this will merge their maps
                    metadata["hyperlinks"].setdefault(orig_row, {}).update(link_map)

        wb.close()

        # 2) Read into pandas
        df = pd.read_excel(path, engine="openpyxl", dtype=str).fillna("")

        # Sanity check: number of DataFrame rows must equal extracted row_indices
        if len(df) != len(row_indices):
            raise ValueError(
                f"Row‐count mismatch for '{path}': "
                f"{len(df)} DataFrame rows vs {len(row_indices)} Excel rows."
            )

        # 3) Attach the original_row_index column
        df["original_row_index"] = row_indices

        # 4) Ensure the ref_column actually exists
        if ref_col not in df.columns:
            raise KeyError(f"Reference column '{ref_col}' not found in '{path}'")

        dfs.append(df)

    return dfs, metadata